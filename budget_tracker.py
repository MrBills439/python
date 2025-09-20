from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Salary Budget Tracker"

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
currency_fmt = '"$"#,##0.00'
percent_fmt = '0.00%'

# Section 1: Overview
overview_headers = ["Month", "Income", "Total Expenses", "Savings (Remaining)", "% Saved"]
ws.append(overview_headers)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Placeholder row for data entry and formulas
ws.append(["September", "", "=SUM(C6:C12)", "=B2-C2", "=D2/B2"])

# Empty row
ws.append([])

# Section 2: Budget & Actual
budget_headers = ["Category", "Budgeted Amount", "Actual Spent", "Over/Under", "% of Income"]
ws.append(budget_headers)

categories = ["Rent", "Groceries", "Transport", "Entertainment", "Utilities", "Subscriptions", "Savings"]

start_row = ws.max_row + 1
for cat in categories:
    row_num = ws.max_row + 1
    ws.append([
        cat,
        "",  # Budgeted
        "",  # Actual
        f"=B{row_num}-C{row_num}",  # Difference
        f"=C{row_num}/$B$2"  # % of Income
    ])

# Total Row
total_row = ws.max_row + 1
ws.append([
    "Total",
    f"=SUM(B{start_row}:{'B'+str(total_row-1)})",
    f"=SUM(C{start_row}:{'C'+str(total_row-1)})",
    f"=B{total_row}-C{total_row}",
    f"=C{total_row}/$B$2"
])

# Header styling for budget section
for cell in ws[start_row - 1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Adjust column widths
for col in ['A', 'B', 'C', 'D', 'E']:
    ws.column_dimensions[col].width = 18

# Number formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for i, cell in enumerate(row[1:4]):  # Budgeted, Actual, Difference
        cell.number_format = currency_fmt
    if isinstance(row[4].value, str) and row[4].value.startswith("="):
        row[4].number_format = percent_fmt

# Add conditional formatting for overspending (if Actual > Budgeted)
ws.conditional_formatting.add(
    f"D{start_row}:{'D'+str(total_row)}",
    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True,
               fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'))
)

# Save file
wb.save("Salary_Budget_Tracker.xlsx")
print("✅ Excel file created: Salary_Budget_Tracker.xlsx")
